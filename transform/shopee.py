import openpyxl

# Abrir a planilha existente
wb = openpyxl.load_workbook(r'C:\Users\User\workspace\gustavo\projetos\shoppe\extract\Shopee_products1.xlsx')
ws = wb['Shopee Products']  # Seleciona a planilha específica

# Definir o intervalo de linhas (ajuste conforme necessário)
for row in range(2, ws.max_row + 1):  # Começar da linha 2 (supondo que a linha 1 é o cabeçalho)
    b_value = ws[f'B{row}'].value
    c_value = ws[f'C{row}'].value

   # Definir o intervalo de linhas (ajuste conforme necessário)
for row in range(2, ws.max_row + 1):  # Começar da linha 2 (supondo que a linha 1 é o cabeçalho)
    b_value = ws[f'B{row}'].value
    c_value = ws[f'C{row}'].value

    # Verificar se os valores estão sendo lidos corretamente
    print(f"Linha {row} - B: '{b_value}', C: '{c_value}'")

    # Certificar-se de que b_value é numérico
    try:
        b_value = float(b_value) if b_value is not None else 0
    except ValueError:
        print(f"Valor não numérico na coluna B, linha {row}. Pulando...")
        continue
    
    # Tratar valor de C (remover o % e converter para float)
    if isinstance(c_value, str):
        c_value = c_value.replace('%', '').strip()  # Remover o símbolo de % e espaços
        try:
            c_value = float(c_value) / 100  # Converter para float e dividir por 100
        except ValueError:
            print(f"Valor não numérico na coluna C, linha {row}. Valor tratado: '{c_value}'")
            continue
    elif c_value is None or c_value == "":
        c_value = 0  # Se o valor for None ou vazio, tratar como 0
    else:
        try:
            c_value = float(c_value) / 100  # Verificar se c_value é numérico e converter
        except ValueError:
            print(f"Valor não numérico na coluna C, linha {row}. Valor tratado: '{c_value}'")
            continue

    # Aplicar a fórmula para a coluna G
    if c_value == 0:
        ws[f'G{row}'] = b_value
        print(f"Coluna G{row} atualizada com valor: {b_value}")
    else:
        ws[f'G{row}'] = b_value / (1 + c_value)
        print(f"Coluna G{row} atualizada com valor: {b_value / (1 + c_value)}")

# Salvar a planilha atualizada
wb.save('C:/Users/User/workspace/gustavo/projetos/shoppe/transform/Shopee_products1_atualizado.xlsx')